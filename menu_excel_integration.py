from tkinter import*
import os
from openpyxl import load_workbook

root = Tk()
root.geometry("1600x700+0+0")
root.title("MASTERMIND FOOD FEST")

Tops = Frame(root,width = 1600,height=50,relief=SUNKEN)
Tops.pack(side=TOP)

f1 = Frame(root,width = 900,height=700,relief=SUNKEN)
f1.pack(side=LEFT)

f2 = Frame(root ,width = 700,height=700,relief=SUNKEN)
f2.pack(side=RIGHT)




lblinfo = Label(Tops, font=( 'aria' ,40, 'bold' ),text="Mastermind Food Fest",fg="violet red",bd=10,anchor='w')
lblinfo.grid(row=0,column=0)
lblinfo = Label(Tops, font=( 'aria' ,20, ),text="-By Hasnat",fg="maroon4",anchor='center')
lblinfo.grid(row=0,column=1)




x = os.getcwd()
filepath="%s/demo.xlsx" %x
wb=load_workbook(filepath)
sheet_rev=wb['Sheet1']
sheet_all=wb['Sheet2']

text_Input=IntVar()

def rev_store():
	sheet_rev['B2'] = Burger_data
	sheet_rev['B3'] = Pizza_data
	sheet_rev['B4'] = Sumi_data
	sheet_rev['B5'] = Doughnut_data
	sheet_rev['B6'] = Set_menu_data
	sheet_rev['B7'] = Sub_data
	sheet_rev['B8'] = Chicken_data
	sheet_rev['B9'] = Fuchka_data
	sheet_rev['B10'] = Cold_coffee_data
	sheet_rev['B11'] = Drink_data
	sheet_rev['B13'] = Revenue
	
	
def food_store(type,c):
	r=3
	e = sheet_all.cell(row=r,column=c)
	j = str(e.value)
	while not j=="None":
		r+=1
		e = sheet_all.cell(row=r,column=c)
		j = str(e.value)
	e.value = type
	
	
def Next():

	food_store(Burger,2)
	food_store(Pizza,3)
	food_store(Sumi,4)
	food_store(Doughnut,5)
	food_store(Set_menu,6)
	food_store(Sub,7)
	food_store(Chicken,8)
	food_store(Fuchka,9)
	food_store(Cold_coffee,10)
	food_store(Drink,11)
	
	food_store(costofburger,13)
	food_store(costofpizza,14)
	food_store(costofsumi,15)
	food_store(costofdoughnut,16)
	food_store(costofset,17)
	food_store(costofsub,18)
	food_store(costofchick,19)
	food_store(costoffuchka,20)
	food_store(costofcoffee,21)
	food_store(costofdrink,22)
	
	food_store(total_cost,24)
	
	wb.save(filepath)



def Clear():
	global Pizza
	global Burger
	global Sumi
	global Doughnut
	global Set_menu
	global Sub
	global Chicken
	global Fuchka
	global Cold_coffee
	global Drink
	global total_cost
	global costofburger
	global costofpizza
	global costofsumi
	global costofdoughnut
	global costofset
	global costofsub
	global costofchick
	global costoffuchka
	global costofcoffee
	global costofdrink
	global Total
	
	Pizza = int()
	Burger = int()
	Sumi = int()
	Doughnut = int()
	Set_menu = int()
	Sub = int()
	Chicken = int()
	Fuchka = int()
	Cold_coffee = int()
	Drink = int()
	Return = int()
	
	costofburger = int()
	costofpizza=int()
	costofsumi = int()
	costofdoughnut = int()
	costofset = int()
	costofsub = int()
	costofchick = int()
	costoffuchka = int()
	costofcoffee = int()
	costofdrink = int()
	
	Paid.set("")
	Total = int()
	
	lblTotal.config(text = Total)
	
	lblReturn.config(text = Return)
	txtBurger.config(text = Burger)
	txtPizza.config(text = Pizza)
	txtSumi.config(text = Sumi)
	txtDoughnut.config(text = Doughnut)
	txtSet_menu.config(text = Set_menu)
	txtSub.config(text = Sub)
	txtChicken.config(text = Chicken)
	txtFuchka.config(text = Fuchka)
	txtCold_coffee.config(text = Cold_coffee)
	txtDrink.config(text = Drink)
	
	lblBurger_cost.config(text = costofburger)
	lblPizza_cost.config(text = costofpizza)
	lblSumi_cost.config(text = costofsumi)
	lblDoughnut_cost.config(text = costofdoughnut)
	lblSet_menu_cost.config(text = costofset)
	lblSub_cost.config(text = costofsub)
	lblChicken_cost.config(text = costofchick)
	lblFuchka_cost.config(text = costoffuchka)
	lblCold_coffee_cost.config(text = costofcoffee)
	lblDrink_cost.config(text = costofdrink)
	
	
	Ref()

def Ref():
	
	global total_cost
	global costofburger
	global costofpizza
	global costofsumi
	global costofdoughnut
	global costofset
	global costofsub
	global costofchick
	global costoffuchka
	global costofcoffee
	global costofdrink
	global Total
	costofburger = Burger*250
	costofpizza=Pizza*100
	costofsumi = Sumi*80
	costofdoughnut = Doughnut*120
	costofset = Set_menu*280
	costofsub = Sub*120
	costofchick = Chicken*100
	costoffuchka = Fuchka*50
	costofcoffee = Cold_coffee*100
	costofdrink = Drink*40

	
	total_cost = int(costofburger + costofpizza + costofsumi + costofdoughnut + costofset + costofsub + costofchick + costoffuchka + costofcoffee + costofdrink)
	Total = "Tk. %d" %(total_cost)
	lblTotal.config(text = Total)
	lblBurger_cost.config(text = costofburger)
	lblPizza_cost.config(text = costofpizza)
	lblSumi_cost.config(text = costofsumi)
	lblDoughnut_cost.config(text = costofdoughnut)
	lblSet_menu_cost.config(text = costofset)
	lblSub_cost.config(text = costofsub)
	lblChicken_cost.config(text = costofchick)
	lblFuchka_cost.config(text = costoffuchka)
	lblCold_coffee_cost.config(text = costofcoffee)
	lblDrink_cost.config(text = costofdrink)
	
	
def Store_data():
   
	global Revenue
	global Burger_data 
	global Pizza_data 
	global Sumi_data 
	global Doughnut_data 
	global Set_menu_data
	global Sub_data
	global Chicken_data
	global Fuchka_data
	global Cold_coffee_data
	global Drink_data
	
	Burger_data += costofburger
	Pizza_data += costofpizza
	Sumi_data+= costofsumi
	Doughnut_data += costofdoughnut
	Set_menu_data += costofset
	Sub_data +=costofsub
	Chicken_data += costofchick
	Fuchka_data += costoffuchka
	Cold_coffee_data += costofcoffee
	Drink_data += costofdrink
	
	Revenue += total_cost
	
	lblstore.config(text = Revenue)

	rev_store()
	Next()
	wb.save(filepath)
	Clear()
	
def Return_money():
	global Return
	pay = int(Paid.get())
	Return = "Tk. %d" %(pay - total_cost)
	lblReturn.config(text = Return)
	
def Burger_add():
	global Burger
	Burger += 1
	txtBurger.config(text = Burger)
	Ref()
	
def Pizza_add():
	global Pizza
	Pizza += 1
	txtPizza.config(text = Pizza)
	
	Ref()
def Sumi_add():
	global Sumi
	Sumi += 1
	txtSumi.config(text = Sumi)
	Ref()
	
def Doughnut_add():
	global Doughnut
	Doughnut += 1
	txtDoughnut.config(text = Doughnut)
	Ref()
	
def Set_menu_add():
	global Set_menu
	Set_menu += 1
	txtSet_menu.config(text = Set_menu)
	Ref()
def Sub_add():
	global Sub
	Sub += 1
	txtSub.config(text = Sub)
	Ref()
def Chicken_add():
	global Chicken
	Chicken += 1
	txtChicken.config(text = Chicken)
	Ref()
def Fuchka_add():
	global Fuchka
	Fuchka += 1
	txtFuchka.config(text = Fuchka)
	Ref()
def Cold_coffee_add():
	global Cold_coffee
	Cold_coffee += 1
	txtCold_coffee.config(text = Cold_coffee)
	Ref()
def Drink_add():
	global Drink
	Drink += 1
	txtDrink.config(text = Drink)
	Ref()


	
def Burger_subtract():
	global Burger
	Burger -= 1
	txtBurger.config(text = Burger)
	Ref()
	
def Pizza_subtract():
	global Pizza
	Pizza -= 1
	txtPizza.config(text = Pizza)
	
	Ref()
def Sumi_subtract():
	global Sumi
	Sumi -= 1
	txtSumi.config(text = Sumi)
	Ref()
def Doughnut_subtract():
	global Doughnut
	Doughnut -= 1
	txtDoughnut.config(text = Doughnut)
	Ref()
def Set_menu_subtract():
	global Set_menu
	Set_menu -= 1
	txtSet_menu.config(text = Set_menu)
	Ref()
def Sub_subtract():
	global Sub
	Sub -= 1
	txtSub.config(text = Sub)
	Ref()
def Chicken_subtract():
	global Chicken
	Chicken -= 1
	txtChicken.config(text = Chicken)
	Ref()
def Fuchka_subtract():
	global Fuchka
	Fuchka -= 1
	txtFuchka.config(text = Fuchka)
	Ref()
def Cold_coffee_subtract():
	global Cold_coffee
	Cold_coffee -= 1
	txtCold_coffee.config(text = Cold_coffee)
	Ref()
def Drink_subtract():
	global Drink
	Drink -= 1
	txtDrink.config(text = Drink)
	Ref()


Pizza = int()
Burger = int()
Sumi = int()
Doughnut = int()
Set_menu = int()
Sub = int()
Chicken = int()
Fuchka = int()
Cold_coffee = int()
Drink = int()

total_cost = int()
Total = int()
Paid = StringVar()
Return = int()

Revenue = int()

Burger_data = int()
Pizza_data = int()
Sumi_data = int()
Doughnut_data = int()
Set_menu_data = int()
Sub_data = int()
Chicken_data = int()
Fuchka_data = int()
Cold_coffee_data = int()
Drink_data = int()
	


costofburger = int()
costofpizza=int()
costofsumi = int()
costofdoughnut = int()
costofset = int()
costofsub = int()
costofchick = int()
costoffuchka = int()
costofcoffee = int()
costofdrink = int()

btnBurger=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Burger", bg="powder blue",command= Burger_add)
btnBurger.grid(row=1, column=1)
txtBurger = Label(f1,font=('ariel' ,16,'bold'), text=Pizza , bd=6,fg="steel blue" ,anchor='center', relief=GROOVE)
txtBurger.grid(row=1,column=2)
btnsubtract_burger=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=2, text="-", bg="powder blue",command= Burger_subtract)
btnsubtract_burger.grid(row=1, column=3)

btnPizza=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Pizza", bg="powder blue",command= Pizza_add)
btnPizza.grid(row=2, column=1)
txtPizza = Label(f1,font=('ariel' ,16,'bold'), text=Pizza , bd=6,fg="steel blue" ,anchor='center',relief=GROOVE )
txtPizza.grid(row=2,column=2)
btnsubtract_pizza=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=2, text="-", bg="powder blue",command= Pizza_subtract)
btnsubtract_pizza.grid(row=2, column=3)

btnSumi=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Sumi", bg="powder blue",command= Sumi_add)
btnSumi.grid(row=3, column=1)
txtSumi = Label(f1,font=('ariel' ,16,'bold'), text=Sumi , bd=6,fg="steel blue" ,anchor='center',relief=GROOVE )
txtSumi.grid(row=3,column=2)
btnsubtract_sumi=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=2, text="-", bg="powder blue",command= Sumi_subtract)
btnsubtract_sumi.grid(row=3, column=3)

btnDoughnut=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Doughnut", bg="powder blue",command= Doughnut_add)
btnDoughnut.grid(row=4, column=1)
txtDoughnut = Label(f1,font=('ariel' ,16,'bold'), text=Doughnut , bd=6,fg="steel blue" ,anchor='center', relief=GROOVE)
txtDoughnut.grid(row=4,column=2)
btnsubtract_doughnut=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=2, text="-", bg="powder blue",command= Doughnut_subtract)
btnsubtract_doughnut.grid(row=4, column=3)

btnSet_menu=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Set menu", bg="powder blue",command= Set_menu_add)
btnSet_menu.grid(row=5, column=1)
txtSet_menu = Label(f1,font=('ariel' ,16,'bold'), text=Set_menu , bd=6,fg="steel blue" ,anchor='center',relief=GROOVE)
txtSet_menu.grid(row=5,column=2)
btnsubtract_set=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=2, text="-", bg="powder blue",command= Set_menu_subtract)
btnsubtract_set.grid(row=5, column=3)

btnSub=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Sub", bg="powder blue",command= Sub_add)
btnSub.grid(row=6, column=1)
txtSub = Label(f1,font=('ariel' ,16,'bold'), text=Sub , bd=6,fg="steel blue" ,anchor='center',relief=GROOVE)
txtSub.grid(row=6,column=2)
btnsubtract_sub=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=2, text="-", bg="powder blue",command= Sub_subtract)
btnsubtract_sub.grid(row=6, column=3)

btnChicken=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Chicken", bg="powder blue",command= Chicken_add)
btnChicken.grid(row=7, column=1)
txtChicken = Label(f1,font=('ariel' ,16,'bold'), text=Chicken , bd=6,fg="steel blue" ,anchor='center',relief=GROOVE)
txtChicken.grid(row=7,column=2)
btnsubtract_chick=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=2, text="-", bg="powder blue",command= Chicken_subtract)
btnsubtract_chick.grid(row=7, column=3)

btnFuchka=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Fuchka", bg="powder blue",command= Fuchka_add)
btnFuchka.grid(row=8, column=1)
txtFuchka = Label(f1,font=('ariel' ,16,'bold'), text=Fuchka , bd=6,fg="steel blue" ,anchor='center',relief=GROOVE)
txtFuchka.grid(row=8,column=2)
btnsubtract_fuchka=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=2, text="-", bg="powder blue",command= Fuchka_subtract)
btnsubtract_fuchka.grid(row=8, column=3)

btnCold_coffee=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Cold coffee", bg="powder blue",command= Cold_coffee_add)
btnCold_coffee.grid(row=9, column=1)
txtCold_coffee = Label(f1,font=('ariel' ,16,'bold'), text=Cold_coffee , bd=6,fg="steel blue" ,anchor='center',relief=GROOVE)
txtCold_coffee.grid(row=9,column=2)
btnsubtract_coffee=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=2, text="-", bg="powder blue",command= Cold_coffee_subtract)
btnsubtract_coffee.grid(row=9, column=3)

btnDrink=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Drink", bg="powder blue",command= Drink_add)
btnDrink.grid(row=10, column=1)
txtDrink = Label(f1,font=('ariel' ,16,'bold'), text=Drink , bd=6,fg="steel blue" ,anchor='center',relief=GROOVE)
txtDrink.grid(row=10,column=2)
btnsubtract_drink=Button(f1,padx=16,pady=2, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=2, text="-", bg="powder blue",command= Drink_subtract)
btnsubtract_drink.grid(row=10, column=3)

lblspace = Label(f1,font=('ariel' ,16,'bold'), text="-----------------------------------" , bd=6,fg="powder blue" ,anchor='e')
lblspace.grid(row=12,column=1)

lblspace = Label(f1,font=('ariel' ,16,'bold'), text="---------------------" , bd=6,fg="white" ,anchor='e')
lblspace.grid(row=12,column=3)

lblspace = Label(f1,font=('ariel' ,16,'bold'), text="---------------------" , bd=6,fg="white" ,anchor='e')
lblspace.grid(row=12,column=4)



btnclear=Button(f1,padx=16,pady=8, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Reset", bg="red",cursor="hand2",command=Clear)
btnclear.grid(row=3, column=5)

btnstore=Button(f1,padx=16,pady=8, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Next", bg="green",command=Store_data)
btnstore.grid(row=5, column=5)
lblstore = Label(f1,font=('ariel' ,16,'bold'), text= 0 , bd=6,fg="steel blue" ,anchor='e')
lblstore.grid(row=7,column=5)





#-------RIFHT SIDE------

lblBurger_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text="Burger",fg="steel blue",bd=10,anchor='w',width=10)
lblBurger_cost.grid(row=1,column=1)
lblBurger_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text=costofburger,fg="steel blue",bd=10,anchor='e')
lblBurger_cost.grid(row=1,column=2)

lblPizza_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text="Pizza",fg="steel blue",bd=10,anchor='w',width=10)
lblPizza_cost.grid(row=2,column=1)
lblPizza_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text=costofpizza,fg="steel blue",bd=10,anchor='e')
lblPizza_cost.grid(row=2,column=2)

lblSumi_cost= Label(f2, font=( 'aria' ,14, 'bold' ),text="Sumi",fg="steel blue",bd=10,anchor='w',width=10)
lblSumi_cost.grid(row=3,column=1)
lblSumi_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text=costofsumi,fg="steel blue",bd=10,anchor='e')
lblSumi_cost.grid(row=3,column=2)

lblDoughnut_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text="Doughnut",fg="steel blue",bd=10,anchor='w',width=10)
lblDoughnut_cost.grid(row=4,column=1)
lblDoughnut_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text=costofdoughnut,fg="steel blue",bd=10,anchor='e')
lblDoughnut_cost.grid(row=4,column=2)

lblSet_menu_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text="Set menu",fg="steel blue",bd=10,anchor='w',width=10)
lblSet_menu_cost.grid(row=5,column=1)
lblSet_menu_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text=costofset,fg="steel blue",bd=10,anchor='e')
lblSet_menu_cost.grid(row=5,column=2)

lblSub_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text="Sub",fg="steel blue",bd=10,anchor='w',width=10)
lblSub_cost.grid(row=6,column=1)
lblSub_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text=costofsub,fg="steel blue",bd=10,anchor='e')
lblSub_cost.grid(row=6,column=2)

lblChicken_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text="Chicken",fg="steel blue",bd=10,anchor='w',width=10)
lblChicken_cost.grid(row=7,column=1)
lblChicken_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text=costofchick,fg="steel blue",bd=10,anchor='e')
lblChicken_cost.grid(row=7,column=2)

lblFuchka_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text="Fuchka",fg="steel blue",bd=10,anchor='w',width=10)
lblFuchka_cost.grid(row=8,column=1)
lblFuchka_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text=costoffuchka,fg="steel blue",bd=10,anchor='e')
lblFuchka_cost.grid(row=8,column=2)

lblCold_coffee_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text="Cold coffee",fg="steel blue",bd=10,anchor='w',width=10)
lblCold_coffee_cost.grid(row=9,column=1)
lblCold_coffee_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text=costofcoffee,fg="steel blue",bd=10,anchor='e')
lblCold_coffee_cost.grid(row=9,column=2)

lblDrink_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text="Drinks",fg="steel blue",bd=10,anchor='w',width=10)
lblDrink_cost.grid(row=10,column=1)
lblDrink_cost = Label(f2, font=( 'aria' ,14, 'bold' ),text=costofdrink,fg="steel blue",bd=10,anchor='e')
lblDrink_cost.grid(row=10,column=2)

lblTotal = Label(f2, font=( 'aria' ,14, 'bold' ),text="Total",fg="green",bd=10,anchor='w',width=10)
lblTotal.grid(row=11,column=1)
lblTotal = Label(f2, font=( 'aria' ,14, 'bold' ),text=Total,fg="steel blue",bd=10,anchor='e')
lblTotal.grid(row=11,column=2)

lblPaid = Label(f2, font=( 'aria' ,14, 'bold' ),text="Paid",fg="blue",bd=10,anchor='w',width=10)
lblPaid.grid(row=12,column=1)
txtPaid = Entry(f2,font=('ariel' ,14,'bold'), textvariable=Paid , bd=6,insertwidth=2,bg="powder blue" ,cursor = "xterm", justify='right')
txtPaid.grid(row=12,column=2)


btnReturn=Button(f2,padx=14,pady=2, bd=10 ,fg="red",font=('ariel' ,14,'bold'),width=10, text="Return", bg="powder blue",command=Return_money)
btnReturn.grid(row=13, column=1)
lblReturn = Label(f2, font=( 'aria' ,14, 'bold' ),text=Return,fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=13,column=2)



root.mainloop()
