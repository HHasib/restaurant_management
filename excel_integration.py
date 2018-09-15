from tkinter import*
import random
import time
from openpyxl import load_workbook


root = Tk()
root.geometry("1600x700+0+0")
root.title("MASTERMIND FOOD FEST")

Tops = Frame(root,bg="white",width = 1600,height=50,relief=SUNKEN)
Tops.pack(side=TOP)

f1 = Frame(root,width = 900,height=700,relief=SUNKEN)
f1.pack(side=LEFT)

f2 = Frame(root ,width = 400,height=700,relief=SUNKEN)
f2.pack(side=RIGHT)

bottom = Frame(root,bg="white",width = 1600,height=20,relief=SUNKEN)
bottom.pack(side=BOTTOM)

filepath="E:/demo.xlsx"
wb=load_workbook(filepath)
sheet=wb.active

text_Input=IntVar()

def food_store(type,c):
	r=1
	e = sheet.cell(row=r,column=c)
	j = str(e.value)
	while not j=="None":
		r+=1
		e = sheet.cell(row=r,column=c)
		j = str(e.value)
	e.value = type
def Next():
	food_store(Burger,1)
	food_store(Pizza,2)
	wb.save(filepath)

def Ref():
	
	global Total
	costofpizza = Pizza*200
	costofburger = Burger*150
	costofsumi = Sumi*50
	costofdoughnut = Doughnut*80
	costofset = Set_menu*300
	costofsub = Sub*100
	costofchick = Chicken*100
	
	total_cost = float(costofburger + costofpizza + costofsumi + costofdoughnut + costofset + costofsub + costofchick)
	Total = total_cost
	lblTotal.config(text = Total)
	pay = float(Paid.get())
	Return = pay - Total
	lblReturn.config(text = Return)
	
	
def Burger_add():
	global Burger
	Burger += 1
	txtBurger.config(text = Burger)
	Ref()
	
def Pizza_add():
	global Pizza
	Pizza += 1
	txtPizza.config(text = Pizza)
	Ref()
def Sumi_add():
	global Sumi
	Sumi += 1
	txtSumi.config(text = Sumi)
	Ref()
def Doughnut_add():
	global Doughnut
	Doughnut += 1
	txtDoughnut.config(text = Doughnut)
	Ref()
def Set_menu_add():
	global Set_menu
	Set_menu += 1
	txtSet_menu.config(text = Set_menu)
	Ref()
def Sub_add():
	global Sub
	Sub += 1
	txtSub.config(text = Sub)
	Ref()
def Chicken_add():
	global Chicken
	Chicken += 1
	txtChicken.config(text = Chicken)
	Ref()

Pizza = int()
Burger = int()
Sumi = int()
Doughnut = int()
Set_menu = int()
Sub = int()
Chicken = int()

Total = int()
Paid = IntVar()
Return = int()



btnBurger=Button(f1,padx=16,pady=8, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Burger", bg="powder blue",command= Burger_add)
btnBurger.grid(row=1, column=0)
txtBurger = Label(f1,font=('ariel' ,16,'bold'), text=Pizza , bd=6,fg="powder blue" ,anchor='e')
txtBurger.grid(row=1,column=3)

btnPizza=Button(f1,padx=16,pady=8, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Pizza", bg="powder blue",command= Pizza_add)
btnPizza.grid(row=2, column=0)
txtPizza = Label(f1,font=('ariel' ,16,'bold'), text=Pizza , bd=6,fg="powder blue" ,anchor='e')
txtPizza.grid(row=2,column=3)

btnSumi=Button(f1,padx=16,pady=8, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Sumi", bg="powder blue",command= Sumi_add)
btnSumi.grid(row=3, column=0)
txtSumi = Label(f1,font=('ariel' ,16,'bold'), text=Sumi , bd=6,fg="powder blue" ,anchor='e')
txtSumi.grid(row=3,column=3)

btnDoughnut=Button(f1,padx=16,pady=8, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Doughnut", bg="powder blue",command= Doughnut_add)
btnDoughnut.grid(row=4, column=0)
txtDoughnut = Label(f1,font=('ariel' ,16,'bold'), text=Doughnut , bd=6,fg="powder blue" ,anchor='e')
txtDoughnut.grid(row=4,column=3)

btnSet_menu=Button(f1,padx=16,pady=8, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Set_menu", bg="powder blue",command= Set_menu_add)
btnSet_menu.grid(row=5, column=0)
txtSet_menu = Label(f1,font=('ariel' ,16,'bold'), text=Set_menu , bd=6,fg="powder blue" ,anchor='e')
txtSet_menu.grid(row=5,column=3)

btnSub=Button(f1,padx=16,pady=8, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Sub", bg="powder blue",command= Sub_add)
btnSub.grid(row=6, column=0)
txtSub = Label(f1,font=('ariel' ,16,'bold'), text=Sub , bd=6,fg="powder blue" ,anchor='e')
txtSub.grid(row=6,column=3)


btnChicken=Button(f1,padx=16,pady=8, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Chicken", bg="powder blue",command= Chicken_add)
btnChicken.grid(row=7, column=0)
txtChicken = Label(f1,font=('ariel' ,16,'bold'), text=Chicken , bd=6,fg="powder blue" ,anchor='e')
txtChicken.grid(row=7,column=3)

txtspace = Label(f1,font=('ariel' ,16,'bold'), text="--------" , bd=6, fg="white", anchor='e')
txtspace.grid(row=8,column=2)

btnTotal=Button(f2,padx=16,pady=8, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="TOTAL", bg="powder blue",command=Ref)
btnTotal.grid(row=12, column=1)

lblTotal = Label(f2, font=( 'aria' ,16, 'bold' ),text=Total,fg="steel blue",bd=10,anchor='e')
lblTotal.grid(row=12,column=2)

btnNext=Button(f1,padx=16,pady=8, bd=10 ,fg="black",font=('ariel' ,16,'bold'),width=10, text="Next", bg="green",command=Next)
btnNext.grid(row=2, column=4)

lblReturn = Label(f2, font=( 'aria' ,16, 'bold' ),text="Return",fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=14,column=1)
lblReturn = Label(f2, font=( 'aria' ,16, 'bold' ),text=Return,fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=14,column=2)

lblPaid = Label(f2, font=( 'aria' ,16, 'bold' ),text="Paid",fg="steel blue",bd=10,anchor='e')
lblPaid.grid(row=13,column=1)
txtPaid = Entry(f2,font=('ariel' ,16,'bold'), textvariable=Paid , bd=6,insertwidth=4,bg="powder blue" ,justify='right')
txtPaid.grid(row=13,column=2)

lblReturn = Label(f2, font=( 'aria' ,12, 'bold' ),text="Return",fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=1,column=2)
lblReturn = Label(f2, font=( 'aria' ,12, 'bold' ),text="Return",fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=2,column=2)
lblReturn = Label(f2, font=( 'aria' ,12, 'bold' ),text="Return",fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=3,column=2)
lblReturn = Label(f2, font=( 'aria' ,12, 'bold' ),text="Return",fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=4,column=2)
lblReturn = Label(f2, font=( 'aria' ,12, 'bold' ),text="Return",fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=5,column=2)
lblReturn = Label(f2, font=( 'aria' ,12, 'bold' ),text="Return",fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=6,column=2)
lblReturn = Label(f2, font=( 'aria' ,12, 'bold' ),text="Return",fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=7,column=2)
lblReturn = Label(f2, font=( 'aria' ,12, 'bold' ),text="Return",fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=8,column=2)
lblReturn = Label(f2, font=( 'aria' ,12, 'bold' ),text="Return",fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=9,column=2)
lblReturn = Label(f2, font=( 'aria' ,12, 'bold' ),text="Return",fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=10,column=2)
lblReturn = Label(f2, font=( 'aria' ,12, 'bold' ),text="Return",fg="steel blue",bd=10,anchor='e')
lblReturn.grid(row=11,column=2)
root.mainloop()